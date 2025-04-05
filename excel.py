import openpyxl


def write_excel_template(filename: str, sheetname: str, listdata: list[list[str]]):
    excel_file = openpyxl.Workbook()
    sheet = excel_file.active
    sheet.column_dimensions["A"].width = 100

    if sheetname != "":
        sheet.title = sheetname

    for item_list in listdata:
        # append는 행을 추가하는 함수, list를 넣어줘야 한다.
        sheet.append(item_list)

    # sheet.cell(row=1, column=1).value = "이름"
    # sheet.cell(row=2, column=1).value = "홍길동"
    excel_file.save(filename)
    excel_file.close()


data_list = [["이름", "나이", "성별"], ["홍길동", 20, "남자"], ["철수", 21, "남자"]]
# print(type(data_list))
write_excel_template("test.xlsx", "첫번째 시트", data_list)


### 엑셀 파일 읽기
excel_file = openpyxl.load_workbook("test.xlsx")
print(excel_file.sheetnames)
sheet = excel_file["첫번째 시트"]
# active_sheet = excel_file.active

for item in sheet.rows:
    print(item[0].value, item[1].value, item[2].value)

excel_file.close()
